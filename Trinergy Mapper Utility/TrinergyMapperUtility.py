#FELIPE IGNACIO CORNEJO ILABACA
#06/02/2020
#V4.0
#PROGRAMA ENCARGADO DE ESCRIBIR EL CÓDIGO DE UN ARCHIVO KML PARA PROYECTO TRINERGY

#Importacion#
import openpyxl
import compresor
import mover as mv
import time

#Def constantes#

NOMBRE = "Trinergy Mapper Utility.xlsx"
HOJA = ["Declarados en Construcción", "Nuevas SE", "Nueva Línea", "Grandes Clientes", "Ampliación", "Nuevos Proyectos", "Electroterminales", "Puertos"]

#Def funciones#

#entrada: nombre del archivo.
#salida: una lista del contenido separado por linea.
#Funcion que lee el archivo y lo hace lista, donde cada elemento es una linea.
def leer(nombre):
    arch = open(nombre, 'r')
    data = arch.read().split('\n')
    tamanyo = len(data)
    arch.close()
    return [data, tamanyo]

#entrada: nombre del archivo y el contenido nuevo.
#salida: entero para revisar que termino el porceso.
#Funcion que escribe un nuevo archivo de texto.
def escribir(nombre, data):
    arch = open(nombre, 'w')
    dataNew = ('\n').join(data)
    arch.write(dataNew)
    arch.close()
    return 1

#entrada: la lista de la data entregada por la funcion *leer*, y el tag a buscar.
#salida: la linea en donde se encuentra el tag que se desea enocontrar.
#Funcion que devuelve el lugar del tag.
def buscarTag(data, tamanyoData, tag):
    txtError = "no se ha encontrado"
    i = 0
    data = list(data)
    while i < tamanyoData:
        if tag in data[i]:
            return i + 1
        i += 1
    return print(txtError)

#entrada: nombre del excel, nombre de la hoja y la linea a revisar.
#salida: una lista de la linea, contiene los datos.
#Funcion que lee y agrega los datos de una linea en especifico a una lista.
def leerExcel(nombre, hoja, columna, fila):
    excelwb = openpyxl.load_workbook(nombre)
    excelpg = excelwb.get_sheet_by_name(hoja)
    lineaF = 1
    while excelpg.cell(column=1, row=lineaF).value is not None:
        lineaF += 1
    columnaF = excelpg.max_column
    linea = []
    while columna <= columnaF:
        dato = excelpg.cell(row = fila, column = columna).value
        linea.append(dato)
        columna += 1
    return [linea, lineaF, columnaF]

#entrada: la lista de la data por lineas y la linea de información del excel
#salida: la data nueva incrustada (NEWDATA)
#funcion que escribe los declarados en construccion y lo integra dentro de la lista data                         
def escribirDeclarados(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- declaradotag -->")
    nombre = str(linea[0])                                                                                            
    empresa = str(linea[1])                                                                                           
    tipo = str(linea[2])                                                                                         
    capacidad = str(linea[3])
    fecha = str(linea[4])
    conexion = str(linea[5])
    x = str(linea[7])
    y = str(linea[6])
    imprimir = linea[8]
    if imprimir == 1:
        if tipo == "Fotovoltaico":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONNECTION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",x,",",y,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "Eolico":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru1</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONNECTION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",x,",",y,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "Diesel":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru2</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONNECTION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",x,",",y,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "Hidroelectrica":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru3</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONNECTION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",x,",",y,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "Geotermica":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru4</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONNECTION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",x,",",y,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "Termosolar":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru5</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONNECTION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",x,",",y,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "Biomasa":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru6</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONNECTION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",x,",",y,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)

        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structDeclarado] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ", fila, " en la hoja Declarados en Construccion")
        return data
    
#entrada: la lista de la data por lineas y la linea de información del excel
#salida: la data nueva incrustada (NEWDATA)
#funcion que escribe los declarados en construccion y lo integra dentro de la lista data                         
def escribirNuevasSE(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- nuevasubtag -->")
    nombre = str(linea[0])                                                                                            
    decreto = str(linea[1])                                                                                           
    transform = str(linea[2])
    estado = str(linea[3])
    x = str(linea[4])
    y = str(linea[5])
    imprimir = linea[6]
    if imprimir == 1:
        structNuevaSE = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#NuevaSE0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevaSE\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DECRETO\">",decreto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TRANSFORM\">",transform,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
        structNuevaSE = ('').join(structNuevaSE)    
        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structNuevaSE] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ", fila , " en la hoja Nuevas SE")
        return data

def escribirGranCliente(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- clientetag -->")
    cliente = str(linea[0])                                                                                            
    barra = str(linea[1])                                                                                           
    suministrador = str(linea[2])
    demanda = str(linea[3])
    x = str(linea[4])
    y = str(linea[5])
    imprimir = linea[6]
    if imprimir == 1:
        structGranCliente = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",cliente,"</name>\n\t\t\t\t\t<styleUrl>#Cliente0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#GranCliente\">\n\t\t\t\t\t\t\t<SimpleData name=\"CLIENTE\">",cliente,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"BARRA\">",barra,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SUMINISTRADOR\">",suministrador,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DEMANDA\">",demanda,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
        structGranCliente = ('').join(structGranCliente)    
        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structGranCliente] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ",fila, " en la hoja Grandes Clientes")
        return data

def escribirElectroterminal(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- electroterminaltag -->")
    cliente = str(linea[0])                                                                                            
    barra = str(linea[1])                                                                                           
    suministrador = str(linea[2])
    demanda = str(linea[3])
    x = str(linea[4])
    y = str(linea[5])
    imprimir = linea[6]
    if imprimir == 1:
        structElectroterminal = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",cliente,"</name>\n\t\t\t\t\t<styleUrl>#Terminal0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#ElectroTerminal\">\n\t\t\t\t\t\t\t<SimpleData name=\"CLIENTE\">",cliente,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"BARRA\">",barra,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SUMINISTRADOR\">",suministrador,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DEMANDA\">",demanda,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
        structElectroterminal = ('').join(structElectroterminal)
        data1 = data[:start]
        data2 = data[start-1:]
        data = data1 + [structElectroterminal] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ",fila, " en la hoja Electroterminales")
        return data

def escribirPuertos(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- puertostag -->")
    nombre = str(linea[0])                                                                                            
    propietario = str(linea[1])                                                                                           
    actividad = str(linea[2])
    calado = str(linea[3])
    x = str(linea[4])
    y = str(linea[5])
    imprimir = linea[6]
    if imprimir == 1:
        structPuerto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Puertos0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Puertos\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PROPIETARIO\">",propietario,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ACTIVIDAD\">",actividad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CALADO\">",calado,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
        structPuerto = ('').join(structPuerto)
        data1 = data[:start]
        data2 = data[start-1:]
        data = data1 + [structPuerto] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ",fila, " en la hoja Electroterminales")
        return data

def escribirAmpliacion(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- ampliaciontag -->")
    obra = str(linea[0])
    decreto = str(linea[1])
    tipo = str(linea[2])
    salida = str(linea[3])
    entrada = str(linea[4])
    estado = str(linea[5])
    x = str(linea[6])
    y = str(linea[7])
    imprimir = linea[8]
    if imprimir == 1:
        if "CNE:" in decreto:
            structAmpliacion = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",obra,"</name>\n\t\t\t\t\t<styleUrl>#Ampliacion0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Ampliacion\">\n\t\t\t\t\t\t\t<SimpleData name=\"OBRA\">",obra,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DECRETO\">",decreto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SALIDA\">",salida,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ENTRADA\">",entrada,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structAmpliacion = ('').join(structAmpliacion)
        else:
            structAmpliacion = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",obra,"</name>\n\t\t\t\t\t<styleUrl>#Ampliacion1</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Ampliacion\">\n\t\t\t\t\t\t\t<SimpleData name=\"OBRA\">",obra,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DECRETO\">",decreto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SALIDA\">",salida,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ENTRADA\">",entrada,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structAmpliacion = ('').join(structAmpliacion)

        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structAmpliacion] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ", fila, " en la hoja: Ampliación")
        return data

def escribirNuevasLineas(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- nuevalineatag -->")
    nombre = str(linea[0])                                                                                            
    decreto = str(linea[1])
    estado = str(linea[2])                                                                                           
    capacidad = str(linea[3])
    trazado = str(linea[4])
    puntos = str(linea[5])
    imprimir = linea[6]
    if imprimir == 1:
        structNuevaLinea = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Linea0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevaLine\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DECRETO\">",decreto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TRAZADO\">",trazado,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<LineString>\n\t\t\t\t\t\t<tessellate>1</tessellate>\n\t\t\t\t\t\t<coordinates>\n\t\t\t\t\t\t\t",puntos,"\n\t\t\t\t\t\t</coordinates>\n\t\t\t\t\t</LineString>\n\t\t\t\t</Placemark>"
        structNuevaLinea = ('').join(structNuevaLinea)    
        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structNuevaLinea] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ", fila, " en la hoja: Nueva Linea")
        return data


def escribirNuevoProyecto(data, tamanyoData, linea, fila):
    start = buscarTag(data, tamanyoData, "<!-- nuevoproyetag -->")
    nombre = str(linea[0])
    estado = str(linea[2])
    empresa = str(linea[3])                                                                                           
    tipo = str(linea[4])
    capacidad = str(linea[5])
    interconexion = str(linea[6])
    punto = str(linea[7])
    rca = str(linea[8])
    aprobacion = str(linea[9])
    x = str(linea[11])
    y = str(linea[12])
    imprimir = linea[13]
    if imprimir == 1:
        if tipo == "Fotovoltaico":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
        elif tipo == "Eolico":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye1</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
        elif tipo == "Diesel" or tipo == "GLP":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye2</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
        elif tipo == "Hidroelectrica":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye3</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
        elif tipo == "Geotermica":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye4</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
        elif tipo == "Termosolar":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye5</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"ESTADO\">",estado,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
        elif tipo == "Biomasa":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye6</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
        elif tipo == "Consumo":
            structNuevoProyecto = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#nproye7</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#NuevoProye\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CAPACIDAD\">",capacidad,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"INTERCONEXION\">",interconexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"PUNTO\">",punto,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"RCA\">",rca,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"APROBACION\">",aprobacion,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structNuevoProyecto = ('').join(structNuevoProyecto)
            
        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structNuevoProyecto] + data2
        return data
    
    else:
        txtError = "Hay algún campo sin rellenar"
        print(txtError, " en la fila: ", fila, " en la hoja: Nuevos Proyectos")
        return data

#Funcion Main, este es el cuerpo del programa parte funcional
#Se usara de exportacion al codigo que contiene la interfaz
#Entrada: none
#Salida: True
def main():
    
    lectura = leer("Base-Utility.kml")#lecutra de kml base
    data = lectura[0]
    tamanyo = lectura[1]
    ######################
    ###NUEVOS PROYECTOS###
    ######################
    print("######################\n###NUEVOS PROYECTOS###\n######################\n\n")
    excelInfo = leerExcel(NOMBRE, HOJA[5], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[5], 1, fila)
        data = escribirNuevoProyecto(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 1/8 completado exitosamente...\n\n")

    ###################
    ###NUEVAS LINEAS###
    ###################
    print("###################\n###NUEVAS LINEAS###\n###################\n\n")
    lectura = leer("Trinergy Utility Mapped.kml")#lecutra de kml base

    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[2], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[2], 1, fila)
        data = escribirNuevasLineas(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir ("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 2/8 completado exitosamente... \n\n")

    ######################
    ###GRANDES CLIENTES###
    ######################
    print("######################\n###GRANDES CLIENTES###\n######################\n\n")
    lectura = leer("Trinergy Utility Mapped.kml")#lecutra de kml base

    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[3], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[3], 1, fila)
        data = escribirGranCliente(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir ("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 3/8 completado exitosamente... \n\n")
    
    #######################
    ###ELECTROTERMINALES###
    #######################
    print("#######################\n###ELECTROTERMINALES###\n#######################\n\n")
    lectura = leer("Trinergy Utility Mapped.kml")#lecutra de kml base

    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[6], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[6], 1, fila)
        data = escribirElectroterminal(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir ("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 4/8 completado exitosamente... \n\n")

    #############
    ###PUERTOS###
    #############
    print("#######################\n###PUERTOS###\n#######################\n\n")
    lectura = leer("Trinergy Utility Mapped.kml")#lecutra de kml base
    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[7], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[7], 1, fila)
        data = escribirPuertos(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir ("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 5/8 completado exitosamente... \n\n")
    ###########################
    ###NUEVAS SUB-ESTACIONES###
    ###########################
    print("###########################\n###NUEVAS SUB-ESTACIONES###\n###########################\n\n")
    lectura = leer("Trinergy Utility Mapped.kml")#lecutra de kml base

    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[1], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas sub estaciones
        excelInfo = leerExcel(NOMBRE, HOJA[1], 1, fila)
        data = escribirNuevasSE(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir ("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 6/8 completado exitosamente... \n\n")

    ##################
    ###AMPLIACIONES###
    ##################
    print("##################\n###AMPLIACIONES###\n##################\n\n")

    lectura = leer("Trinergy Utility Mapped.kml")#lecutra de kml base

    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[4], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[4], 1, fila)
        data = escribirAmpliacion(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir ("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 7/8 completado exitosamente... \n\n")
    ################################
    ###DECLARADOS EN CONSTRUCCIÓN###
    ################################
    print("################################\n###DECLARADOS EN CONSTRUCCION###\n################################\n\n")
    lectura = leer("Trinergy Utility Mapped.kml")#lecutra de kml base

    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[0],1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1
    while fila <= totalLines: #Aqui empieza la escritura de Declarados
        excelInfo = leerExcel(NOMBRE, HOJA[0], 1, fila)
        data = escribirDeclarados(data, tamanyo, excelInfo[0], fila)
        fila += 1

    pase = escribir ("Trinergy Utility Mapped.kml", data)

    print("\n\nProceso 8/8 completado exitosamente... \n\n")
    compresor.main()
    time.sleep(30)
    mv.mover()
    return True
