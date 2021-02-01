#FELIPE IGNACIO CORNEJO ILABACA
#06/02/2020
#V4.0
#PROGRAMA ENCARGADO DE ESCRIBIR EL CÓDIGO DE UN ARCHIVO KML PARA PROYECTO TRINERGY

#Importacion#
import openpyxl
from progress.spinner import Spinner
import time
from os import system
import compresor
import mover

#Def constantes#

NOMBRE = "Trinergy Mapper PMGD.xlsx"
HOJA = ["Declarados en Construcción", "Actuales"]

#entrada: nombre del archivo.
#salida: una lista del contenido separado por linea.
#Funcion que lee el archivo y lo hace lista, donde cada elemento es una linea.
def leer(nombre):
    arch = open(nombre, 'r')
    data = arch.read().split('\n')
    tamanyo = len(data)
    arch.close()
    return [data, tamanyo]

#entrada: nombre del archivo y el contenido nuevo.
#salida: entero para revisar que termino el porceso.
#Funcion que escribe un nuevo archivo de texto.
def escribir(nombre, data):
    arch = open(nombre, 'w')
    dataNew = ('\n').join(data)
    arch.write(dataNew)
    arch.close()
    return 1

#entrada: la lista de la data entregada por la funcion *leer*, y el tag a buscar.
#salida: la linea en donde se encuentra el tag que se desea enocontrar.
#Funcion que devuelve el lugar del tag.
def buscarTag(data, tamanyoData, tag):
    txtError = "no se ha encontrado"
    i = 0
    data = list(data)
    while i < tamanyoData:
        if tag in data[i]:
            return i + 1
        i += 1
    return print(txtError)

#entrada: nombre del excel, nombre de la hoja y la linea a revisar.
#salida: una lista de la linea, contiene los datos.
#Funcion que lee y agrega los datos de una linea en especifico a una lista.
def leerExcel(nombre, hoja, columna, fila):
    excelwb = openpyxl.load_workbook(nombre)
    excelpg = excelwb.get_sheet_by_name(hoja)
    lineaF = 1
    while excelpg.cell(column=1, row=lineaF).value is not None:
        lineaF += 1
    columnaF = excelpg.max_column
    linea = []
    while columna <= columnaF:
        dato = excelpg.cell(row = fila, column = columna).value
        linea.append(dato)
        columna += 1
    return [linea, lineaF, columnaF]
    
def escribirOperativos0(data, tamanyoData, linea):
    start = buscarTag(data, tamanyoData, "<!-- operativetag -->")
    nombre = str(linea[0])                                                                                            
    empresa = str(linea[1])                                                                                           
    tipopmg = str(linea[2])
    tipotecno = str(linea[3])
    potencia = str(linea[4])
    region = str(linea[5])
    comuna = str(linea[6])
    conexion = str(linea[7])
    se = str(linea[8])
    fecha = str(linea[9])
    x = str(linea[10])
    y = str(linea[11])
    imprimir = linea[12]
    if imprimir == 1:
        if tipotecno == "PMGD Fotovoltaico":
            structOperative = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Operative0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Operative\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipopmg,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TECNO\">",tipotecno,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"COMUNA\">",comuna,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structOperative = ('').join(structOperative)
        elif tipotecno == "PMGD Hidroelectrica":
            structOperative = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Operative1</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Operative\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipopmg,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TECNO\">",tipotecno,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"COMUNA\">",comuna,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structOperative = ('').join(structOperative)
        elif tipotecno == "PMGD Eolico":
            structOperative = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Operative2</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Operative\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipopmg,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TECNO\">",tipotecno,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"COMUNA\">",comuna,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structOperative = ('').join(structOperative)
        elif tipotecno == "PMGD Geotermica":
            structOperative = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Operative3</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Operative\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipopmg,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TECNO\">",tipotecno,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"COMUNA\">",comuna,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structOperative = ('').join(structOperative)
        elif tipotecno == "PMGD GLP" or tipotecno == "PMGD Diesel":
            structOperative = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Operative4</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Operative\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipopmg,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TECNO\">",tipotecno,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"COMUNA\">",comuna,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structOperative = ('').join(structOperative)
        else:
            structOperative = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#Operative4</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#Operative\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipopmg,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TECNO\">",tipotecno,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"COMUNA\">",comuna,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structOperative = ('').join(structOperative)
        
        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structOperative] + data2
        return data
        
    else:
        return data


def escribirDeclarado0(data, tamanyoData, linea):
    start = buscarTag(data, tamanyoData, "<!-- declaradotag -->")
    nombre = str(linea[0])                                                                                            
    empresa = str(linea[1])
    fecha = str(linea[2])
    tipo = str(linea[3])
    potencia = str(linea[4])
    region = str(linea[5])
    conexion = str(linea[6])
    se = str(linea[7])
    datos = str(linea[8])
    x = str(linea[9])
    y = str(linea[10])
    imprimir = linea[11]
    if imprimir == 1:
        if tipo == "PMGD Fotovoltaico":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru0</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DATOS\">",datos,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "PMGD Hidroelectrica":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru1</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DATOS\">",datos,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "PMGD Eolico":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru2</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DATOS\">",datos,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "PMGD Geotermica":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru3</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DATOS\">",datos,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        elif tipo == "PMGD GLP" or tipo == "PMGD Diesel":
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru4</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DATOS\">",datos,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        else:
            structDeclarado = "\n\t\t\t\t<Placemark>\n\t\t\t\t\t<name>",nombre,"</name>\n\t\t\t\t\t<styleUrl>#DecConstru4</styleUrl>\n\t\t\t\t\t<ExtendedData>\n\t\t\t\t\t\t<SchemaData schemaUrl=\"#DecConstru\">\n\t\t\t\t\t\t\t<SimpleData name=\"NOMBRE\">",nombre,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"EMPRESA\">",empresa,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"FECHA\">",fecha,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"TIPO\">",tipo,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"POTENCIA\">",potencia,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"REGION\">",region,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"CONEXION\">",conexion,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"SE\">",se,"</SimpleData>\n\t\t\t\t\t\t\t<SimpleData name=\"DATOS\">",datos,"</SimpleData>\n\t\t\t\t\t\t</SchemaData>\n\t\t\t\t\t</ExtendedData>\n\t\t\t\t\t<Point>\n\t\t\t\t\t\t<gx:drawOrder>1</gx:drawOrder>\n\t\t\t\t\t\t<coordinates>",y,",",x,",0</coordinates>\n\t\t\t\t\t</Point>\n\t\t\t\t</Placemark>"
            structDeclarado = ('').join(structDeclarado)
        
        data1 = data[:start]
        data2 = data[start - 1:]
        data = data1 + [structDeclarado] + data2
        return data
    
    else:
        return data

def main():
    spinner = Spinner('Leyendo: ')
    ##################
    ####Operativos####
    ##################

    lectura = leer("Base-PMGD.kml")
    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[1], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[1], 1, fila)
        data = escribirOperativos0(data, tamanyo, excelInfo[0])
        fila += 1
        spinner.next()
    pase = escribir("Trinergy PMGD Mapped.kml", data)

    #################################
    ####Declarado en Construcción####
    #################################

    lectura = leer("Trinergy PMGD Mapped.kml")
    data = lectura[0]
    tamanyo = lectura[1]

    excelInfo = leerExcel(NOMBRE, HOJA[0], 1, 1)

    totalLines = excelInfo[1]
    totalColumns = excelInfo[2]

    fila = 1

    while fila <= totalLines: #Aqui empieza la escritura de nuevas LINEAS
        excelInfo = leerExcel(NOMBRE, HOJA[0], 1, fila)
        data = escribirDeclarado0(data, tamanyo, excelInfo[0])
        fila += 1
        spinner.next()
        
    pase = escribir("Trinergy PMGD Mapped.kml", data)

    compresor.main()
    time.sleep(30)
    mover.mover()



